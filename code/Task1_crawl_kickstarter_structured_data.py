import requests, time
import openpyxl, json, re

wb = openpyxl.Workbook()
s1 = wb.active
row = 2

file_title = 'title	blurb	金额单位	金额标志	總目標	已認繳	地区	发起者	发起者支持数量	发起者发起数量	state	发起日期	截止日期	持续时间	分类	支持者	留言数量	背景故事文本	背景故事图片数量	视频链接	详情页链接'
for index,value in enumerate(file_title.split('\t')):
    s1.cell(row=1,column=index+1,value=value)


try:
    with open('done', 'r', encoding='utf-8')as file_r:
        done_list = file_r.read().split('\n')
except:
    done_list = []
while '' in done_list:
    done_list.remove('')


cookie = 'vis=d2b1dc29b99ca012-379c91ee7e5018a5-219e8a1304b5a64fv1; ajs_anonymous_id=d2b1dc29b99ca012-379c91ee7e5018a5-219e8a1304b5a64fv1; _ga=GA1.2.1248420906.1672208447; _pxvid=c23a0fe3-8677-11ed-a52d-4b4164646349; __ssid=ec845dc01c3c6a910f4292e7f58dc94; __stripe_mid=f665616d-f4f2-41b2-bc92-c3d355654b155ec4f2; _fbp=fb.1.1672208700701.2135200776; lang=zh; optimizely_current_variations=%7B%7D; pxcts=8343ead0-880d-11ed-86e1-56457767584e; woe_id=gWWIeiudHky%2BfGa4%2B3wC8VOBfu05nBiwFg%2F7Fnjs%2B4iNYt0Z46YKO8KA6hpMRVywlIqk7xtCTgi%2B%2Bo%2BsJsSIgqMa8e%2BAgDxugxuaSHBio2Um37gWytM%2BXg%3D%3D--nZESw8ydKPIvS4ZN--sSIPKoKCWVNZMC2MhZp%2FfQ%3D%3D; _gid=GA1.2.485518844.1672748816; last_page=https%3A%2F%2Fwww.kickstarter.com%2Fprojects%2Ftentaclemag%2Fcrossworlds-the-tabletop-comic-book-wargame%3Fref%3Ddiscovery_category; _pxhd=DYB9n9VTbMQ4hVsTLwGBkSWiQiXcBwFc9uDx0t6tTXHk5Eghyi3sXQKgGmtU0U4KP8KT2zatzQ9HLNZsfogAXQ==:QsguXWlcgFMgpbeYL6g3/mHzF8jBkjUYvJOX84glcGcMSHoXdQF1MfQZ5q/t9D6VxphsZrC3kvB/6qXDgqH6v/s3H4RtkeMpCFpeTIXIo0Spc3gKVv98YvfPOyYQoR0gh28cXSbMveqoKa8L/9olRw==; __cf_bm=i0os3Pv0ajx0se3OGs5e6Qi5JwxfmN314.SBb6TaVZg-1672794711-0-AR25z7QNO0J1cnQcSE+2nwEDE0OXB5LjNW5l0JweejTAMOFmTpM8CyFpX5xzaCKtGkYpCgSUivfi3vErRoF4t2g=; local_offset=2116; _pxff_cc=U2FtZVNpdGU9TGF4Ow; _ksr_session=o8D23TiiVnchWVu84hvpNhDGHe95e46SkMLGU%2BrLSTaF%2FqL7kef%2BBtp41c96yIkbFa8A2jWxZ2lulOpF5nnvcOulnhWfTpkEg0zO6CIYyZYjPollvM9f3ysM6uiRuJwzAindNpEfJUSwoOctLjtneqIp2swLY281204PSGJ1PoIQA3iExO3d3SSuhzRGStHwyjZW0NwEVE453uGT1yqebIjCGOrVfyK5Vw1duC6IIbc6%2FA%2BjOE2hitL6xjBmVkMj2IXaWnR%2F76oaF16sniIP3xEuN1M%3D--dtYXMk4MsBbi%2F6NT--l%2BTqv2iy37LOAEcu%2FPKeLg%3D%3D; request_time=Wed%2C+04+Jan+2023+01%3A11%3A57+-0000; _px2=eyJ1IjoiYzNiMzRiNTAtOGJjYy0xMWVkLTk5MWItNWI5ODJlN2ZiYjk5IiwidiI6ImMyM2EwZmUzLTg2NzctMTFlZC1hNTJkLTRiNDE2NDY0NjM0OSIsInQiOjE2NzI3OTUwMTc0MDgsImgiOiI0NWU4MDEyYWEyOTAzODM3YjZiYjRmZjU3YmM3NTU3MWY0MGYwZTNjOGYwNDM1ZjI1NGQ0YWQ2M2Y2ZTJiMmFmIn0; __stripe_sid=fc91127a-e0d0-4de5-a3f3-31773036ef68575286'
headers_of_list_page = {
    'accept': 'application/json, text/javascript, */*; q=0.01',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'zh-CN,zh;q=0.9',
    'cookie': cookie,
    'referer': 'https://www.kickstarter.com/discover/advanced?sort=magic&seed=2787038&page=1',
    'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
    'x-csrf-token': 'pU8hjDD0aBIfgifWJ2TuCB6BA60LBPxC1bSn0kcphvCKOEzO6Mep6nCVoj112i9Lfaakxeo7Q8IutEoW-UIlLQ',
    'x-requested-with': 'XMLHttpRequest'

}

headers_background_story = {
    'accept': '*/*',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
    'content-length': '1135',
    'content-type': 'application/json',
    'cookie': 'vis=0879bca3f28643fa-dd17ae3580aacec7-721994c7bfc74c8ev1; lang=zh; woe_id=7UXQBbG1CoSXf%2BUbRNhGH%2FxY0xV0SPsQkOS5lIvQ0hvDcaKGmYZ4K5QPnooThBQEflAgvG6ThnBR%2FwU%2B3b2x3dZ4HHHxotboCERuBYs0tlUSZvbP9kNGcg%3D%3D--2z5W5cs94jYpETV9--7KWp6gzp1moXszH2cBT3vQ%3D%3D; pxcts=b3162500-867d-11ed-b916-416455454f70; _pxvid=b12b7b80-867d-11ed-88ff-465153474645; __stripe_mid=73e90338-e165-4cbb-a2a0-d1c2dde6e1c2d95bb6; ajs_anonymous_id=0879bca3f28643fa-dd17ae3580aacec7-721994c7bfc74c8ev1; _pxhd=m4TDFZa2Oxave46lgvoaAO3PI9ffnCr2m3bOmNWR8h1OU4AeyGaBssU2bsHvhzfLTCR43saZUd-XbXj2nDscAw==:tbWWWVJ2/Gmgwwcy4PPHWELexwKG4PYDCzAT9pynUTsu3ep7TabiEOcMwMcjvHAQHrVzulLx7hVWYESjiw-Loh4Rh9LjAfEa/v2VzSwRKFs=; _ga=GA1.2.2077699921.1672285361; _gid=GA1.2.1700703496.1672285361; __ssid=e5a3440391849652d2a1b0b9c759377; optimizely_current_variations=%7B%7D; _px2=eyJ1IjoiOWYwM2JlMDAtODc0MS0xMWVkLTk2ZTEtYWIzNGU4OWM1NGE2IiwidiI6ImIxMmI3YjgwLTg2N2QtMTFlZC04OGZmLTQ2NTE1MzQ3NDY0NSIsInQiOjE2NzIyOTU0NjI4NTMsImgiOiJkMGZjMmQzZGRlYTBjZDQyNjE2Yzk0MmIyNGViNjRlYzU3ZGQ2MTM5NzFiMzc5MWM2NjNkZDIwMGY5OTZhNzE5In0=; __cf_bm=fJkk7rthbio0.rsjOdk70fed4YamFziVwh7aAqV0yHs-1672295170-0-AVSO2o4LsBqG76fbbWzLri1u/tjwlaUIRUAdUwpSc1dsH74BT/gVHNJDwaP0UmmO0A4E9lsvG2bLZcN7T+6Gpmg=; last_page=https%3A%2F%2Fwww.kickstarter.com%2Fprojects%2Fthornwillow%2Fjames-joyces-ulysses-the-centennial-edition%3Fref%3Ddiscovery; _ksr_session=lDT874Mge28nAxeUwTW4dAotAyyd4lI9G3J5hINVh1d9CFr0ASUKdIELQtl7vCIuThH24Xe7vRqe1YVYi%2FrDv1nf9Bu1hN9Io%2BGqLAV%2B1OpdvNSyG8dV0%2BgTK7MyURbskR2u4KArRc3aDs%2FBZySflyZHYsazaz88eGSslqpMwV%2FOKP98I2h%2FagHCFDS%2B%2BTafPPByHbTvSWWR%2FoGZVR0z2LAG%2BA%2FlYhOEwsO0LuoHKYK6Ew7%2FDCdI0mBNCsly6WFFPM2iButY6TkKSfqkhB7g%2FR1U1v8%3D--RctsTUVRz4R5BsVd--UbrWzovUG7Ahd5MGjvzLcg%3D%3D; request_time=Thu%2C+29+Dec+2022+06%3A26%3A27+-0000; local_offset=-30389',
    'origin': 'https://www.kickstarter.com',
    'referer': 'https://www.kickstarter.com/projects/thornwillow/james-joyces-ulysses-the-centennial-edition?ref=discovery',
    'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Microsoft Edge";v="108"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54',
    'x-csrf-token': 'oZx1gtU2me3L1MYHs8NcBNjOw2cQIo4u3vmxRmqJ8dtTEBpDQrC6tq9SKLZSaOvbsT5RNNSk3vRk8GJB3DUN1A'
}

headers_pj_stats_url = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'zh-CN,zh;q=0.9',
    'cache-control': 'max-age=0',
    # 'cookie': 'vis=d2b1dc29b99ca012-379c91ee7e5018a5-219e8a1304b5a64fv1; lang=zh; woe_id=e0dA5hs%2FrY3vUG3%2BXxeay7fWfTYRUrcmxsYC0rixv6Sn7xqi8qoEQbEsBx2fFBvj8knP91HwYJSVTZxz1Dn3nlEtjemj3G0WFc13EhqenKJnnrHr5VQFDg%3D%3D--vgKVg8ep2GY9SwPY--hk2ClcjM1tX%2F7QcIvFKUiw%3D%3D; optimizely_current_variations=%7B%7D; ajs_anonymous_id=d2b1dc29b99ca012-379c91ee7e5018a5-219e8a1304b5a64fv1; _ga=GA1.2.1248420906.1672208447; _gid=GA1.2.1211049520.1672208447; pxcts=ca75eb87-8677-11ed-b05a-67586a635462; _pxvid=c23a0fe3-8677-11ed-a52d-4b4164646349; __ssid=ec845dc01c3c6a910f4292e7f58dc94; __stripe_mid=f665616d-f4f2-41b2-bc92-c3d355654b155ec4f2; _fbp=fb.1.1672208700701.2135200776; __stripe_sid=de963f10-e590-41fb-a2e3-cd41e3c744b4a8fca7; _pxhd=-C2zScEKIwj0Nw71Xwrxnir6N5NyZKqeSNchDEiOwFMxBMQCG3uPq3/uJ2s4FfVe83v4/3V9IQI917OvwxDkSg==:A6C-U0//oZbQ/ncXsCzoTqkyORiE3FvIA3DwKM5c/vsZSV4XQZBe1YhPcuXt1EGwBKLz8Gj6F6tamNhESwpNHc0d82HEGzrsLR2-i4SPnsaxx1Fdp4/cHRwFEyDDWZDglViGhpXDibtviEe6pKuR7g==; __cf_bm=mrXWd.RUAImDKatBaKk6WxEcvap4E6sVwsTtDyNVo64-1672225029-0-AQg9V81C2STgKPfdNFvby9OJz2/7nHLO0fFH0qED0z9SyXzPuo+qpmahmxQeITuuwPE1hhLlx491eNNkRpH5RjU=; last_page=https%3A%2F%2Fwww.kickstarter.com%2Fprofile%2Fnilawriters; local_offset=3707; _pxff_cc=U2FtZVNpdGU9TGF4Ow; _ksr_session=FrOcwIL4vU6ZYA1zWwPIkErzCoQohTrmU%2BkPdWyjekQpKBuBeyLfQVuv4oxz90WInBp8ulxMqgOxuZJELhui9eQ%2BdLCaZlqyF13W8DVFw4ZU9WthIhphK%2FEr%2FF6PK1ObXEUn13%2F3X0PFdKLblQy5IdLA2Vd%2Fyb%2FRHtJJoyNZl8Kb0wJk5GFcNejTXreL%2Bjp6gOS9X1OT0k5AC%2FVSR%2FqowZ98iJB7nWXfgJQhbaB5x1I6AV2JepnZu%2BrA%2B0lu96tzG%2B4QRzaP589DWa%2FJcUVBbrduWSA%3D--jLOkx8Ti2wGr6nOP--8wk2ZDNZyeMggkLE1TitIA%3D%3D; request_time=Wed%2C+28+Dec+2022+11%3A11%3A39+-0000; _px2=eyJ1IjoiNjI4Yzg4ZTAtODZhMC0xMWVkLWJlOWEtOGJhYzJkMmU5MWJkIiwidiI6ImMyM2EwZmUzLTg2NzctMTFlZC1hNTJkLTRiNDE2NDY0NjM0OSIsInQiOjE2NzIyMjYxOTkyOTcsImgiOiIwMDUzZTgzZjFiZGM2NDE2NTEyZTg0ZjIyNzBhYTNiYjIxMjY1NzY5NDUwZTRmMzFjMzZjMGZkOTg2NWJlOWQ3In0',
    'cookie': cookie,

    'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'same-origin',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36'
}

headers_ = {
    'accept': 'application/json, text/javascript, */*; q=0.01',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'zh-CN,zh;q=0.9',
    # 'cookie': 'vis=0879bca3f28643fa-dd17ae3580aacec7-721994c7bfc74c8ev1; lang=zh; woe_id=7UXQBbG1CoSXf%2BUbRNhGH%2FxY0xV0SPsQkOS5lIvQ0hvDcaKGmYZ4K5QPnooThBQEflAgvG6ThnBR%2FwU%2B3b2x3dZ4HHHxotboCERuBYs0tlUSZvbP9kNGcg%3D%3D--2z5W5cs94jYpETV9--7KWp6gzp1moXszH2cBT3vQ%3D%3D; optimizely_current_variations=%7B%7D; pxcts=b3162500-867d-11ed-b916-416455454f70; _pxvid=b12b7b80-867d-11ed-88ff-465153474645; __stripe_mid=73e90338-e165-4cbb-a2a0-d1c2dde6e1c2d95bb6; ajs_anonymous_id=0879bca3f28643fa-dd17ae3580aacec7-721994c7bfc74c8ev1; _pxhd=YNE375RCBxmdy3iThjecMFUxV4A5KHiw4UR6ANnWB-QdarDdDh1kf6iLSjSf04ZV34jWxjwm99x5Z5fHXx--wA==:In693RA1gme3YTeqbr2IGAidjxKUc1KncFBLvdhI-011UKjucH1UDfDDV-LFnWQHK59ME12OrGkiDwVSg9VXxCfYBC3XMDuXuwOc/y0DvhZfTiUGpm9ccORgcu-rL6ja4xt/8HJOofEeqUyCee17fQ==; local_offset=3201; request_time=Wed%2C+28+Dec+2022+07%3A54%3A24+-0000; _ksr_session=9lyW%2FMzQoVQ3JRH%2FOu4Cc%2BVDqIWJEsthQONllPZd6BP7Z0Uke8Ou4VPpu0FV3CjwqXxexWccPe5iCOF0DJA9CQA4wh8bbStEBsYdhqT0ONJ%2FSwy7XmQj79cGVoDpzTqOB0uSxRQ2yrVRLj%2BGU%2BwjazGgCw0not4AzY4PMVoEwcM7taGFbjOOGKOH3c46uLxqG46K0TpMlvk4OhbchaZE5IUUt9uzPK9qYJT9qk%2Ft6pPHkC5kMFkql0SjR9h7gYw5EjIbGaazDXQCYpBoCEYUYcxi%2Bu8%3D--iXvAx8Z52KvKzcHn--%2BidIZb0XbDX2bPcLQTuqDA%3D%3D; __cf_bm=pHt8NbrBSWw3J0luFsS0SPXgS3vfRLRdmSgXAlhCYiI-1672217529-0-ASBYuvADb/7ToBLIlCayDSr5Pz9kPAk78HLhj3/VKjenWEWmZGPc1+zl70yBxkI/YushjfAWlyMGMeOjkOfWmhQ=; _px2=eyJ1IjoiZDRiYmI1YjAtODY4NC0xMWVkLTgwYzEtN2QzMjFiYjEwMDc4IiwidiI6ImIxMmI3YjgwLTg2N2QtMTFlZC04OGZmLTQ2NTE1MzQ3NDY0NSIsInQiOjE2NzIyMTg2MjIxMTEsImgiOiIyMGRiYTZmMmI2MzU3YTEwMzkzOTY2MmFlNjhhZmJmOGQ5ZjY4YTcyZTc2MTRjNjY0YTE0NzBmYjRiNjg3NTllIn0',
    'cookie': cookie,

    'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Google Chrome";v="108"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'empty',
    'sec-fetch-mode': 'cors',
    'sec-fetch-site': 'same-origin',
    # 'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
    'x-csrf-token': 'a__-K9faNTv4xFCxyrlnTzAHuXcFvbJo-IQZ1Wl-3kiZc5HqQFwWYJxCvgArEtCQWfcrJME74rJCjcrS38IiRw'
}

headers_detail_Video = {
    'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
    'accept-encoding': 'gzip, deflate, br',
    'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
    'cache-control': 'max-age=0',
    'cookie': cookie,
    'sec-ch-ua': '"Not?A_Brand";v="8", "Chromium";v="108", "Microsoft Edge";v="108"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'sec-fetch-dest': 'document',
    'sec-fetch-mode': 'navigate',
    'sec-fetch-site': 'none',
    'sec-fetch-user': '?1',
    'upgrade-insecure-requests': '1',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36 Edg/108.0.1462.54'
}


def replace_all_tags(data):
    # src = re.findall(' src="([\d\D]*?)"',data)
    # len_of_src = len(src)

    for i in re.findall('<[\d\D]*?>',data):
        # if ' src="' in i:
        #     src = re.search(' src="([\d\D]*?)"',i).group(1)
        #     data = data.replace(i, src)
        # else:
        #     data = data.replace(i, '')
        data = data.replace(i, '')
    return data
count_of_data = 0


def date_to_stamp(date):
    timeArray = time.strptime(date, "%Y-%m-%d %H:%M:%S")
    timeStamp = int(time.mktime(timeArray))
    return timeStamp


def lastTime(t):
    s = ""
    days_s = ""
    hour_s = ""
    min_s = ""
    sec = t

    while True:
        if sec <= 60:
            second_s = str(sec)
            if days_s != "":
                s = days_s + "天"
            if hour_s != "":
                s = s + hour_s + "小时"
            if min_s != "":
                s = s + min_s + "分钟"
            s = s + second_s + "秒"
            return s
        if sec >= 60 and sec < 3600:
            minutes = sec // 60
            min_s = str(minutes)
            if sec % 60 == 0:
                if days_s != "":
                    s = days_s + "天"
                if hour_s != "":
                    s = s + hour_s + "小时"
                if min_s != "":
                    s = s + min_s + "分钟"
                return s
            else:
                sec = sec % 60
        if sec >= 3600 and sec < 86400:
            # global sec, hour_s
            hours = sec // 3600
            hour_s = str(hours)

            if sec % 3600 == 0:

                if days_s != "":
                    s = days_s + "天"
                if hour_s != "":
                    s = s + hour_s + "小时"
                return s
            else:
                sec = sec % 3600
        if sec >= 86400:
            days = sec // 86400
            days_s = str(days)
            if sec % 86400 == 0:

                return days
            else:
                sec = sec % 86400
            # print('The number of minutes is {0:.2f}'.format(days))


def count_date(st, et):
    st_as_tss = date_to_stamp(st)
    et_as_tss = date_to_stamp(et)
    time_c = et_as_tss - st_as_tss
    time_c_as_str2 = str(lastTime(time_c))

    if '天' not in time_c_as_str2:
        time_c_as_str2 += '天'

    return time_c_as_str2


for page in range(1, 501):

    url = 'https://www.kickstarter.com/discover/advanced?google_chrome_workaround&state=successful&sort=magic&seed=2787038&page={}'.format(
        page)
    # 
    url = 'https://www.kickstarter.com/discover/advanced?google_chrome_workaround&state=successful&category_id=14&sort=magic&seed=2787903&page={}'.format(page)
    time_of_req_bef = time.time()
    try:
        resp = requests.get(url, headers=headers_of_list_page, timeout=20).json()
    except:
        time.sleep(5)
        resp = requests.get(url, headers=headers_of_list_page, timeout=20).json()


    time_of_req_spend = time.time() - time_of_req_bef
    print('ListPage请求花费的时间:', time_of_req_spend)
    range_time_s_one_page = time.time()

    dlist = resp['projects']
    for i in dlist:
        range_time_s = time.time()


        backers_count = i['backers_count']
        blurb = i['blurb']
        category = i['category']
        converted_pledged_amount = i['converted_pledged_amount']
        country = i['country']
        country_displayable_name = i['country_displayable_name']
        created_at_times = i['created_at']
        creator = i['creator']
        currency = i['currency']
        currency_symbol = i['currency_symbol']
        currency_trailing_code = i['currency_trailing_code']
        current_currency = i['current_currency']
        deadline = i['deadline']

        disable_communication = i['disable_communication']
        fx_rate = i['fx_rate']
        goal = i['goal']
        the_id = int(i['id'])
        is_starrable = i['is_starrable']
        launched_at = i['launched_at']
        name = i['name']
        photo = i['photo']
        pledged = i['pledged']
        profile = i['profile']
        slug = i['slug']

        spotlight = i['spotlight']
        staff_pick = i['staff_pick']
        state = i['state']
        if state == 'failed' or state == 'successful':
            pass

        else:
            print( name, state, '非成功或失败')
            continue

        state_changed_at = i['state_changed_at']
        static_usd_rate = i['static_usd_rate']
        urls = i['urls']
        project_url = urls['web']['project']


        if project_url in done_list:
            print('此条爬过',name,project_url)
            continue
        else:
            done_list.append(project_url)

        usd_exchange_rate = i['usd_exchange_rate']
        usd_pledged = i['usd_pledged']
        usd_type = i['usd_type']

        try:
            location = i['location']
            location_short_name = location['short_name']
        except:
            location_short_name = ''
        category_slug = category['slug']
        creator_name = creator['name']
        creator_url = creator['urls']['web']['user']

        category_slug = category['slug']

        category = str(category)
        creator = str(creator)
        # location = str(location)
        photo = str(photo)
        profile = str(profile)
        urls = str(urls)
        # 背景故事部分
        slug_of_bgst = project_url.split('?')[0].replace('https://www.kickstarter.com/projects/', '')

        form_data = [
            {"operationName": "Campaign",
             "variables": {"slug": slug_of_bgst},
             "query": "query Campaign($slug: String!) {\n  project(slug: $slug) {\n    id\n    isSharingProjectBudget\n    risks\n    story(assetWidth: 680)\n    currency\n    spreadsheet {\n      displayMode\n      public\n      url\n      data {\n        name\n        value\n        phase\n        rowNum\n        __typename\n      }\n      dataLastUpdatedAt\n      __typename\n    }\n    environmentalCommitments {\n      id\n      commitmentCategory\n      description\n      __typename\n    }\n    __typename\n  }\n}\n"},
            {"operationName": "FetchProjectSignalAndWatchStatus",
             "variables": {"pid": the_id},
             "query": "query FetchProjectSignalAndWatchStatus($pid: Int) {\n  project(pid: $pid) {\n    ...project\n    __typename\n  }\n  me {\n    ...user\n    __typename\n  }\n}\n\nfragment project on Project {\n  id\n  pid\n  isDisliked\n  isLiked\n  isWatched\n  isWatchable\n  isLaunched\n  __typename\n}\n\nfragment user on User {\n  id\n  uid\n  canSeeConfirmWatchModal\n  canSeeConfirmSignalModal\n  isEmailVerified\n  __typename\n}\n"}]
        form_data = json.dumps(form_data)

        back_ground_st_url = 'https://www.kickstarter.com/graph'

        headers_background_story['referer'] = project_url

        st = time.time()
        resp_of_back_ground_st = requests.post(back_ground_st_url, headers=headers_background_story, data=form_data)
        try:
            resp_of_back_ground_st = resp_of_back_ground_st.json()
        except:
            print([resp_of_back_ground_st.text])
            exit()

        print('spend_time_back_ground_st_url', time.time() - st)
        back_ground_story = resp_of_back_ground_st[0]['data']['project']['story']
        src = re.findall(' src="([\d\D]*?)"', back_ground_story)
        len_of_src = len(src)

        back_ground_story = replace_all_tags(back_ground_story)
        while '\n\n' in back_ground_story:
            back_ground_story = back_ground_story.replace('\n\n', '\n')


        time_of_req_creator = time.time()
        try:
            resp_of_creator_url = requests.get(creator_url, headers=headers_pj_stats_url, timeout=80).text

        except:
            time.sleep(20)
            resp_of_creator_url = requests.get(creator_url, headers=headers_pj_stats_url, timeout=80).text
        print('spend_time_of_req_creator', time.time() - time_of_req_creator)
        try:
            spoed = re.search('已支持[\d\D]*?</span>', resp_of_creator_url).group()
            spoed = spoed.replace('<span class="count">', '').replace('</span>', '').replace('\n', '')
        except:
            spoed = ''
        try:
            sended = re.search('已發起[\d\D]*?</span>', resp_of_creator_url).group()
            sended = sended.replace('<span class="count">', '').replace('</span>', '').replace('\n', '')
        except:
            sended = ''


        time_of_pj_stats_url =time.time()
        # pj_stats_url = project_url.replace('?ref=discovery', '/stats.json?v=1')
        pj_stats_url = project_url.split('?ref=discovery')[0]+'/stats.json?v=1'

        try:
            resp_pj = requests.get(pj_stats_url, headers=headers_,  timeout=70).json()

            try:

                resp_pj = requests.get(pj_stats_url, headers=headers_,  timeout=70).json()

            except:
                time.sleep(20)
                resp_pj = requests.get(pj_stats_url, headers=headers_,  timeout=80).json()

            comments_count = resp_pj['project']['comments_count']
            pledged_now = resp_pj['project']['pledged']
            backers_count_now = resp_pj['project']['backers_count']
            state_now = resp_pj['project']['state']
        except:
            pledged_now = pledged
            backers_count_now = backers_count
            state_now = state
            comments_count = ''

        print('spend_time_time_of_pj_stats_url', time.time() - time_of_pj_stats_url)

        timeArray_launched_at = time.localtime(launched_at)
        launched_at = time.strftime("%Y-%m-%d %H:%M:%S", timeArray_launched_at)

        timeArray_deadline = time.localtime(deadline)
        deadline = time.strftime("%Y-%m-%d %H:%M:%S", timeArray_deadline)

        # 视频

        bf_req_ts = time.time()
        try:
            resp = requests.get(project_url, headers=headers_detail_Video).text
        except:
            time.sleep(20)
            resp = requests.get(project_url, headers=headers_detail_Video).text

        print('spend_time_Video_url', time.time() - bf_req_ts)

        video_path = re.search('<div class="video-player" id="video_pitch" data-video-url="(.*?)"', resp)

        if video_path:
            video_path = video_path.group(1)
        else:
            video_path = ''

        duration = count_date(launched_at,deadline)


        result2 = [
            # 名称
            name,
            # 简述
            blurb,
            # 金额单位
            currency,
            # 金额单位符号
            currency_symbol,
            # 总目标
            goal,
            # 目前筹集金额
            pledged_now,
            # 地区
            location_short_name,
            # 发起人
            creator_name,
            # 发起人支持
            spoed,
            # 发起人发起
            sended,
            # 状态
            state,
            # 开始时间
            launched_at,
            # 截止时间
            deadline,
            # 持续时间
            duration,
            # source_url,
            category_slug,
            # 支持者
            backers_count_now,
            # 留言
            comments_count,
            # 背景故事文字
            back_ground_story,
            # 背景故事图片数量
            len_of_src,
            # 视频链接
            video_path,
            # 链接
            project_url
        ]
        count_of_data += 1
        print(video_path)
        print('本次已采集：', count_of_data, '条数据','现在页数:',page, result2)
        print('一轮循环花费的时间:', time.time() - range_time_s)

        print()
        with open('done', 'a', encoding='utf-8')as file_w:
            file_w.write(project_url)
            file_w.write('\n')

        save_file_name = 'result_kickstarter'

        for index, value in enumerate(result2):
            s1.cell(row=row, column=index + 1, value=value)
        row += 1
        wb.save(f'{save_file_name}.xlsx')
        exit()

    print('一页花费的时间:', time.time() - range_time_s_one_page)

wb.save(f'{save_file_name}.xlsx')

